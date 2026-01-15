# routes/placa_routes.py
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    current_app,
    send_file,   # ✅ NUEVO
)
from flask_login import login_required, current_user

from models.base import db
from models.placa import Placa
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


placa_bp = Blueprint("placa_bp", __name__)


# -----------------------------------------------------------
# Listar placas
# -----------------------------------------------------------
@placa_bp.route("/", methods=["GET"])
@login_required
def listar_placas():
    try:
        placas = Placa.query.order_by(Placa.fecha_registro.desc()).all()
        return render_template("placas.html", placas=placas)
    except Exception as e:
        current_app.logger.exception(f"Error al listar placas: {e}")
        flash("Ocurrió un error al cargar las placas.", "danger")
        return render_template("placas.html", placas=[])


# -----------------------------------------------------------
# Registrar nueva placa
# -----------------------------------------------------------
@placa_bp.route("/nueva", methods=["POST"])
@login_required
def nueva_placa():
    try:
        numero_placa = (request.form.get("numero_placa") or "").strip().upper()
        propietario = (request.form.get("propietario") or "").strip() or None
        color_cabezal = (request.form.get("color_cabezal") or "").strip() or None

        # ✅ NUEVO (opcional): permitir cargar identificador fijo desde el form si lo agregás
        identificador_fijo = (request.form.get("identificador_fijo") or "").strip() or None

        if not numero_placa:
            flash("Debe ingresar un número de placa.", "warning")
            return redirect(url_for("placa_bp.listar_placas"))

        # Verificar duplicados
        if Placa.query.filter_by(numero_placa=numero_placa).first():
            flash("Esta placa ya está registrada.", "danger")
            return redirect(url_for("placa_bp.listar_placas"))

        nueva = Placa(
            numero_placa=numero_placa,
            propietario=propietario,
            color_cabezal=color_cabezal,
            identificador_fijo=identificador_fijo,  # ✅ NUEVO
            usuario_id=current_user.id if current_user else None,
        )

        db.session.add(nueva)
        db.session.commit()

        flash(f"Placa {numero_placa} agregada exitosamente.", "success")
        return redirect(url_for("placa_bp.listar_placas"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error al agregar placa: {e}")
        flash(f"Error al registrar la placa: {str(e)}", "danger")
        return redirect(url_for("placa_bp.listar_placas"))


# -----------------------------------------------------------
# Actualizar datos de placa (propietario y color)
# -----------------------------------------------------------
@placa_bp.route("/actualizar/<int:placa_id>", methods=["POST"])
@login_required
def actualizar_placa(placa_id):
    try:
        placa = Placa.query.get_or_404(placa_id)

        propietario = (request.form.get("propietario") or "").strip() or None
        color_cabezal = (request.form.get("color_cabezal") or "").strip() or None

        # ✅ NUEVO (opcional): si algún día actualizás individualmente
        identificador_fijo = (request.form.get("identificador_fijo") or "").strip() or None

        placa.propietario = propietario
        placa.color_cabezal = color_cabezal
        placa.identificador_fijo = identificador_fijo  # ✅ NUEVO

        db.session.commit()

        flash(f"Placa {placa.numero_placa} actualizada correctamente.", "success")
        return redirect(url_for("placa_bp.listar_placas"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error al actualizar placa: {e}")
        flash(f"Error al actualizar la placa: {str(e)}", "danger")
        return redirect(url_for("placa_bp.listar_placas"))


# -----------------------------------------------------------
# Cambiar estado (Activa / Inactiva)
# -----------------------------------------------------------
@placa_bp.route("/estado/<int:placa_id>", methods=["POST"])
@login_required
def cambiar_estado(placa_id):
    try:
        placa = Placa.query.get_or_404(placa_id)
        nuevo_estado = (request.form.get("estado") or "").strip()

        if nuevo_estado not in ["Activa", "Inactiva"]:
            flash("Estado inválido.", "warning")
            return redirect(url_for("placa_bp.listar_placas"))

        placa.estado = nuevo_estado
        db.session.commit()

        flash(
            f"Estado de la placa {placa.numero_placa} cambiado a {nuevo_estado}.",
            "info",
        )
        return redirect(url_for("placa_bp.listar_placas"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error al cambiar estado de placa: {e}")
        flash(f"Error al actualizar el estado de la placa: {str(e)}", "danger")
        return redirect(url_for("placa_bp.listar_placas"))


# -----------------------------------------------------------
# ✅ Guardar cambios en lote (un solo Guardar para toda la tabla)
# -----------------------------------------------------------
@placa_bp.route("/actualizar_batch", methods=["POST"])
@login_required
def actualizar_placas_batch():
    try:
        updates = {}

        # request.form trae keys tipo:
        # propietario[26], color_cabezal[26], estado[26], identificador_fijo[26]
        for key, value in request.form.items():
            if key.startswith("propietario["):
                placa_id = int(key.split("[", 1)[1].split("]")[0])
                updates.setdefault(placa_id, {})["propietario"] = value.strip() or None

            elif key.startswith("color_cabezal["):
                placa_id = int(key.split("[", 1)[1].split("]")[0])
                updates.setdefault(placa_id, {})["color_cabezal"] = value.strip() or None

            elif key.startswith("estado["):
                placa_id = int(key.split("[", 1)[1].split("]")[0])
                estado = value.strip()
                if estado in ["Activa", "Inactiva"]:
                    updates.setdefault(placa_id, {})["estado"] = estado

            # ✅ NUEVO: identificador fijo
            elif key.startswith("identificador_fijo["):
                placa_id = int(key.split("[", 1)[1].split("]")[0])
                updates.setdefault(placa_id, {})["identificador_fijo"] = value.strip() or None

        if not updates:
            flash("No hay cambios para guardar.", "info")
            return redirect(url_for("placa_bp.listar_placas"))

        # Traer todas las placas involucradas en una sola consulta
        ids = list(updates.keys())
        placas = Placa.query.filter(Placa.id.in_(ids)).all()
        placas_por_id = {p.id: p for p in placas}

        # Aplicar cambios
        for pid, data in updates.items():
            p = placas_por_id.get(pid)
            if not p:
                continue

            if "propietario" in data:
                p.propietario = data["propietario"]

            if "color_cabezal" in data:
                p.color_cabezal = data["color_cabezal"]

            if "estado" in data:
                p.estado = data["estado"]

            # ✅ NUEVO
            if "identificador_fijo" in data:
                p.identificador_fijo = data["identificador_fijo"]

        db.session.commit()

        flash("✅ Cambios guardados correctamente.", "success")
        return redirect(url_for("placa_bp.listar_placas"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error en actualizar_placas_batch: {e}")
        flash(f"❌ Error al guardar cambios: {str(e)}", "danger")
        return redirect(url_for("placa_bp.listar_placas"))
    
# -----------------------------------------------------------
# ✅ Exportar placas a Excel (XLSX)
# -----------------------------------------------------------
@placa_bp.route("/export/excel", methods=["GET"])
@login_required
def exportar_placas_excel():
    try:
        placas = Placa.query.order_by(Placa.fecha_registro.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Placas"

        headers = [
            "ID",
            "Placa",
            "Propietario",
            "Color Cabezal",
            "Identificador Fijo",
            "Estado",
            "Fecha Registro",
            "Usuario ID",
        ]
        ws.append(headers)

        # Estilo encabezados
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        for p in placas:
            fecha = getattr(p, "fecha_registro", None)
            if fecha:
                # Si viene con tz o raro, lo dejamos como string bonito
                try:
                    fecha = fecha.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    fecha = str(fecha)

            ws.append([
                p.id,
                p.numero_placa,
                p.propietario or "",
                p.color_cabezal or "",
                getattr(p, "identificador_fijo", None) or "",
                getattr(p, "estado", None) or "",
                fecha or "",
                getattr(p, "usuario_id", None) or "",
            ])

        # Auto-ajustar columnas (con límite para que no se haga gigante)
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_len = 12
            for cell in ws[col_letter]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"placas_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        current_app.logger.exception(f"Error al exportar placas a Excel: {e}")
        flash("❌ Ocurrió un error exportando a Excel.", "danger")
        return redirect(url_for("placa_bp.listar_placas"))